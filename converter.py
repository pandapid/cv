# -*- coding: utf-8 -*-
"""
Pustaka konversi tabel ⇄ VCF (vCard 3.0) tanpa pandas.
Mendukung: CSV/TXT/TSV, Excel (XLSX via openpyxl), dan VCF.
"""
import csv
import os
import re
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import load_workbook, Workbook

# ---------- Util umum ----------

def detect_delimiter(sample_path: str, fallback: str = ",") -> str:
    try:
        with open(sample_path, "r", encoding="utf-8", errors="ignore") as f:
            sample = f.read(4096)
        # deteksi sederhana
        candidates = ["	", ";", ",", "|", ":"]
        counts = {d: sample.count(d) for d in candidates}
        return max(counts, key=counts.get) if max(counts.values()) > 0 else fallback
    except Exception:
        return fallback

# ---------- vCard helpers ----------

def _escape_vcard_value(value: str) -> str:
    value = value.replace("\", "\\").replace(";", "\;").replace(",", "\,").replace("
", "\n")
    return value

def _unfold_vcard_lines(lines: List[str]) -> List[str]:
    out = []
    for line in lines:
        if line.startswith(" ") or line.startswith("	"):
            if out:
                out[-1] += line.lstrip()
        else:
            out.append(line)
    return out

def build_vcard_row(row: Dict[str, Any]) -> str:
    given = str(row.get("given_name", "") or "").strip()
    family = str(row.get("family_name", "") or "").strip()
    full_name = str(row.get("full_name", "") or "").strip()
    if not full_name:
        full_name = " ".join([p for p in [given, family] if p]).strip() or "Tanpa Nama"

    org = str(row.get("org", "") or "").strip()
    title = str(row.get("title", "") or "").strip()

    street = str(row.get("street", "") or "").strip()
    city = str(row.get("city", "") or "").strip()
    region = str(row.get("region", "") or "").strip()
    postal = str(row.get("postal", "") or "").strip()
    country = str(row.get("country", "") or "").strip()

    note = str(row.get("note", "") or "").strip()

    phones: List[Tuple[str, str]] = []
    def add_phone(val: Any, typ: str):
        num = str(val or "").strip()
        if num:
            phones.append((typ, num))
    add_phone(row.get("phone"), "VOICE")
    add_phone(row.get("phone_mobile"), "CELL")
    add_phone(row.get("phone_home"), "HOME")
    add_phone(row.get("phone_work"), "WORK")

    for k, v in list(row.items()):
        if k.startswith("phone_") and k not in {"phone_mobile", "phone_home", "phone_work"}:
            label = k.split("_", 1)[1].upper() if "_" in k else "VOICE"
            add_phone(v, label)

    emails: List[str] = []
    for key in ["email", "email_alt"]:
        val = str(row.get(key, "") or "").strip()
        if val:
            emails.append(val)

    v = [
        "BEGIN:VCARD",
        "VERSION:3.0",
        f"N:{_escape_vcard_value(family)};{_escape_vcard_value(given)};;;",
        f"FN:{_escape_vcard_value(full_name)}",
    ]
    for typ, num in phones:
        v.append(f"TEL;TYPE={typ}:{_escape_vcard_value(num)}")
    for em in emails:
        v.append(f"EMAIL;TYPE=INTERNET:{_escape_vcard_value(em)}")
    if org:
        v.append(f"ORG:{_escape_vcard_value(org)}")
    if title:
        v.append(f"TITLE:{_escape_vcard_value(title)}")
    if any([street, city, region, postal, country]):
        adr = ["", "", street, city, region, postal, country]
        v.append("ADR;TYPE=HOME:" + ";".join(_escape_vcard_value(x) for x in adr))
    if note:
        v.append(f"NOTE:{_escape_vcard_value(note)}")
    v.append("END:VCARD")
    return "
".join(v)

TEL_RE = re.compile(r"^TEL(?:;TYPE=([^:;]+))?:(.+)$", re.IGNORECASE)
EMAIL_RE = re.compile(r"^EMAIL(?::|;TYPE=[^:]+:)(.+)$", re.IGNORECASE)
N_RE = re.compile(r"^N:(.*)$", re.IGNORECASE)
FN_RE = re.compile(r"^FN:(.*)$", re.IGNORECASE)
ORG_RE = re.compile(r"^ORG:(.*)$", re.IGNORECASE)
TITLE_RE = re.compile(r"^TITLE:(.*)$", re.IGNORECASE)
ADR_RE = re.compile(r"^ADR(?:;TYPE=[^:]+)?:([^$]+)$", re.IGNORECASE)
NOTE_RE = re.compile(r"^NOTE:(.*)$", re.IGNORECASE)

def parse_vcf(text: str) -> List[Dict[str, Any]]:
    cards = []
    blocks = text.split("BEGIN:VCARD")
    for block in blocks:
        if "END:VCARD" not in block:
            continue
        lines = [x.strip("
") for x in block.splitlines() if x.strip()]
        lines = _unfold_vcard_lines(lines)
        rec: Dict[str, Any] = {
            "full_name": "", "given_name": "", "family_name": "",
            "phones": [], "emails": [],
            "org": "", "title": "",
            "street": "", "city": "", "region": "", "postal": "", "country": "",
            "note": "",
        }
        for line in lines:
            m = FN_RE.match(line)
            if m:
                rec["full_name"] = m.group(1).replace("\,", ",").replace("\;", ";").replace("\n", "
").replace("\\", "\"); continue
            m = N_RE.match(line)
            if m:
                parts = m.group(1).split(";")
                family = parts[0] if len(parts) > 0 else ""
                given = parts[1] if len(parts) > 1 else ""
                rec["family_name"] = family.replace("\,", ",").replace("\;", ";")
                rec["given_name"] = given.replace("\,", ",").replace("\;", ";"); continue
            m = TEL_RE.match(line)
            if m:
                _type = (m.group(1) or "VOICE").upper(); number = m.group(2)
                rec.setdefault("phones_typed", []).append((_type, number))
                rec["phones"].append(number); continue
            m = EMAIL_RE.match(line)
            if m:
                rec["emails"].append(m.group(1)); continue
            m = ORG_RE.match(line)
            if m:
                rec["org"] = m.group(1); continue
            m = TITLE_RE.match(line)
            if m:
                rec["title"] = m.group(1); continue
            m = ADR_RE.match(line)
            if m:
                adr = m.group(1).split(";")
                rec["street"] = adr[2] if len(adr) > 2 else ""
                rec["city"] = adr[3] if len(adr) > 3 else ""
                rec["region"] = adr[4] if len(adr) > 4 else ""
                rec["postal"] = adr[5] if len(adr) > 5 else ""
                rec["country"] = adr[6] if len(adr) > 6 else ""; continue
            m = NOTE_RE.match(line)
            if m:
                rec["note"] = m.group(1); continue
        if not rec["full_name"]:
            rec["full_name"] = (rec.get("given_name", "") + " " + rec.get("family_name", "")).strip() or "Tanpa Nama"
        cards.append(rec)
    return cards

# ---------- IO tabel ----------

def load_table(path: str, delimiter: Optional[str] = None) -> List[Dict[str, Any]]:
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx"]:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h) if h is not None else "" for h in rows[0]] if rows else []
        data = []
        for r in rows[1:]:
            rec = {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
            data.append(rec)
        wb.close()
        return data
    elif ext in [".csv", ".txt", ".tsv"]:
        delim = delimiter or ("	" if ext == ".tsv" else detect_delimiter(path))
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.DictReader(f, delimiter=delim)
            return list(reader)
    else:
        raise ValueError("Ekstensi input tidak didukung. Gunakan csv/txt/tsv/xlsx.")


def save_table(rows: List[Dict[str, Any]], path: str):
    ext = os.path.splitext(path)[1].lower()
    # konsolidasi header
    headers: List[str] = []
    for r in rows:
        for k in r.keys():
            if k not in headers:
                headers.append(k)
    if ext in [".csv", ".txt", ".tsv"]:
        delim = "	" if ext == ".tsv" else ","
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers, delimiter=delim)
            writer.writeheader()
            for r in rows:
                writer.writerow(r)
    elif ext in [".xlsx"]:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        wb.save(path)
        wb.close()
    else:
        raise ValueError("Ekstensi output tidak didukung. Gunakan csv/txt/tsv/xlsx.")

# ---------- Konversi utama ----------

def table_to_vcf(input_path: str, output_path: str, delimiter: Optional[str] = None):
    rows = load_table(input_path, delimiter=delimiter)
    vcards = [build_vcard_row(r) for r in rows]
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("
".join(vcards) + ("
" if vcards else ""))


def vcf_to_table(input_path: str, output_path: str):
    with open(input_path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()
    records = parse_vcf(text)
    norm = []
    for r in records:
        o = dict(r)
        o["phones"] = "; ".join(r.get("phones", []))
        o["emails"] = "; ".join(r.get("emails", []))
        norm.append(o)
    save_table(norm, output_path)
